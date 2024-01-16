import pandas as pd
import numpy as np
from openpyxl import load_workbook
from os.path import exists as file_exists
import glob

print('\nCombine all Excell files in current folder,',
      'then concaenate multiple "Sales text" lines into one.\n',
      'Expects all files have same structure with columns:\n',
      '\tmaterial Number, Language, Material description, Client level deleted,',
      'Sales org. level deleted, Sales_text.\n')

out_file = 'product_names.xlsx'
files_in_folder = glob.glob('*.xls?')

def read_in_file(in_file, in_SheetName = '', header_row = 0):
    
    try:
        if in_SheetName == '':
            print (f"Loading first sheet in file {in_file}")
            df = pd.read_excel(in_file, sheet_name = 0, header = header_row)
        else:
            df = pd.read_excel(in_file, in_SheetName, header = header_row)
        column_names = rename_columns(df.columns.values)
        df.columns = column_names
    
    except TypeError:
        if in_SheetName == '':
            print (f"Loading first sheet in file {in_file}")
            wb = load_workbook(filename = in_file, data_only = True)
            df = pd.DataFrame(wb[wb.sheetnames[0]].values)
            
        else:
            wb = load_workbook(filename = in_file, data_only = True)
            df = pd.DataFrame(wb[in_SheetName].values)

        wb.close()
        column_names = rename_columns(df.iloc[header_row].values)
        df.columns = column_names
        df.drop(index = list(np.arange(header_row+1)), inplace = True)
        df.reset_index(drop = True, inplace = True)
    
    except FileNotFoundError:
        df = pd.DataFrame()
        print(f'File {in_file} not found')
    
    except Exception:
        raise 

    return df

def rename_columns(column_names):
    for item in enumerate(column_names):
        try:
            column_names[item[0]] = '_'.join(item[1].split())
        except:
            column_names[item[0]] = item[1]
    return column_names

def strip_led_zeros(i):
    try:
        i = str(int(i))
    except:
        i
    return i

for i in files_in_folder:
    if i != out_file:
        if i == files_in_folder[0]:
            df = read_in_file(i) 
        else:
            df = pd.concat([df, read_in_file(i)], axis = 0, ignore_index = True)

columns = df.columns.to_list()

lang_to_keep = '8'
print(f"Filter only records with '{lang_to_keep}' in column '{columns[1]}'")
records = df.shape[0]
df = df[df.loc[:, columns[1]] == lang_to_keep]
records -= df.shape[0]
print(f'{records} removed.')

df_duplicates = df[df.duplicated(subset = df.columns[0], keep = False)]
df = df[~df.duplicated(subset = df.columns[0], keep = False)]

df_duplicates = df_duplicates.replace(np.nan, '')
concat_list = list()
for material in df_duplicates.iloc[:,0].unique():
    df_transposed = df_duplicates[df_duplicates.iloc[:,0] == material].transpose()
    df_transposed['Concat'] = df_transposed.apply(''.join, axis = 1)
    df_transposed.iloc[0,-1] = df_transposed.iloc[0,0]
    df_transposed.iloc[1,-1] = df_transposed.iloc[1,0]
    concat_list.append(df_transposed.iloc[:,-1])

df_duplicates = pd.DataFrame(concat_list)
df_duplicates = df_duplicates.replace('', np.nan)

df = pd.concat([df, df_duplicates], axis=0, ignore_index=True)

print(f"\nRemoving leading zeros from '{columns[0]}'")
df.loc[:, columns[0]] = df.loc[:, columns[0]].apply(
    lambda x: strip_led_zeros(x))

markers_to_delete = df.loc[:, columns[4]].value_counts().index.to_list()
records = df.shape[0]
print(f"Remove records with {markers_to_delete} in column '{columns[4]}'")
df = df[df.iloc[:, 4].isna()]
records -= df.shape[0]
print(f'{records} removed.')

df = df.iloc[:, [0, 2, 5]]

if file_exists(out_file):
    print(f'\nReplacing {out_file}\n')
else:
    print(f'\nCreating {out_file}\n')
    
with pd.ExcelWriter(out_file, mode = 'w') as writer:
    df.to_excel(writer, sheet_name = 'Total', index=False)
    df_duplicates.to_excel(
        writer, sheet_name='Transformed_duplicates', index=False)
    
    # writer.save()
print(
    f"{df_duplicates.shape[0]} records with multiple line description found and concatenated.")
input('Press any key to finish.')